# -*-coding:utf-8-*-

# 标准库
import time
import sys
import codecs
import os
import json
import io
# 第三方库
import requests
from huzhifeng import dumpObj, hasKeys
from openpyxl import load_workbook
from openpyxl import Workbook

# Set default encoding to utf-8
reload(sys)
sys.setdefaultencoding('utf-8')
# requests.packages.urllib3.disable_warnings()

# 全局变量
PWD = sys.path[0]
SAVE_PATH = PWD + '/result/'
RET_OK = 0
RET_ERR = -1
MAX_TRIES = 3
MAX_DAYS = 60
stationNameCodeMap = {}
stationCodeNameMap = {}
startTimeMap = {}
arriveTimeMap = {}

EXCEL_SHEET_NUMBER = 24
cityList = []
START_EXCEL_NAME = 'start.xlsx'
ARRIVE_EXCEL_NAME = 'arrive.xlsx'
SUM_EXCEL_NAME = 'sum.xlsx'
startWB = Workbook()
arriveWB = Workbook()
sumWB = Workbook()

def queryTickets(queryDate, from_station_code, to_station_code):
    time.sleep(1)
    parameters = [
        ('leftTicketDTO.train_date', queryDate),
        ('leftTicketDTO.from_station', from_station_code),
        ('leftTicketDTO.to_station', to_station_code),
        ('purpose_codes', "ADULT"),
    ]
    headers = {
        'Accept-Encoding': 'gzip, deflate, sdch, br',
        'Accept-Language': 'zh-CN,zh;q=0.8,en-US;q=0.6,en;q=0.4,zh-TW;q=0.2',
        'Cache-Control':'no-cache',
        #'Host': 'kyfw.12306.cn',
        'Referer': 'https://kyfw.12306.cn/otn/leftTicket/init',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36'
    }
    try:
        r = requests.get('https://kyfw.12306.cn/otn/leftTicket/queryA', params=parameters, headers=headers,
              timeout=10, verify=False)
        while r.url == "http://www.12306.cn/mormhweb/logFiles/error.html":
            r = requests.get('https://kyfw.12306.cn/otn/leftTicket/queryA', params=parameters, headers=headers,
                             timeout=10, verify=False)
        print r.url
    except (requests.exceptions.RequestException, requests.exceptions.ConnectionError):
        try:
            print(u'网络问题，等它三秒试试')
            time.sleep(3)
            r = requests.get('https://kyfw.12306.cn/otn/leftTicket/queryA', params=parameters,
                      timeout=8, verify=False)
        except (requests.exceptions.RequestException, requests.exceptions.ConnectionError) as e:
            print e

            fromCity = stationCodeNameMap[from_station_code]
            toCity = stationCodeNameMap[to_station_code]
            saveData = {'fromCity': u'' + fromCity,
                        'toCity': u'' + toCity,
                        'queryDate': queryDate}
            with io.open(PWD + '/config.json', 'w', encoding='utf-8') as outfile:
                outfile.write(unicode(json.dumps(saveData, ensure_ascii=False)))
            outfile.close()
            startWB.save(SAVE_PATH + START_EXCEL_NAME)
            arriveWB.save(SAVE_PATH + ARRIVE_EXCEL_NAME)
            sumWB.save(SAVE_PATH + SUM_EXCEL_NAME)
            print "休息23秒，准备重新启动程序"
            time.sleep(23)
            start()
            return RET_OK

    print(">>>>>>>>>>>>>>>>>>>>>")
    obj = r.json()
    if hasKeys(obj, ['status', 'httpstatus', 'data']):
        return obj
    else:
        print(u'4查询失败')
        print obj
        if hasKeys(obj, ['messages']):
            dumpObj(obj['messages'])
        return RET_ERR


# 根据cityLookUp初始化stationNameCodeMap 和 stationCodeNameMap
def initStation():
    f = codecs.open(PWD + "/cityLookUp.txt", "r")
    data = f.readline()
    f.close()
    station_list = data.split('@')
    if len(station_list) < 1:
        print(u'站点数据库初始化失败, 数据异常')
        return None
    station_list = station_list[1:]
    for station in station_list:
        items = station.split('|')  # bji|北京|BJP|beijing|bj|2
        if len(items) < 5:
            print(u'忽略无效站点: %s' % (items))
            continue
        stationNameCodeMap[items[1].decode('utf-8')] = items[2]
        stationCodeNameMap[items[2]] = items[1]
    return stationNameCodeMap


def getStationByName(name):
    if stationNameCodeMap.has_key(name):
        return stationNameCodeMap[name]
    else:
        print name + "没找到"
        return 0


def initCitiesList():
    f = open(PWD + "/cities.txt", "r")
    global cityList
    cityList = f.readlines()
    f.close()


def start():
    print(u"hello，运行中不要打开result目录中的start.xls和arrive.xls，如果打开了现在关闭还来得及")
    initStation()
    initCitiesList()
    initialWB()
    with open(PWD + '/config.json', 'r') as f:
        crashSave = json.load(f)
    startCityFind = False
    toCityFind = False
    queryDate = crashSave['queryDate']

    print queryDate
    for fromCityIndex in range(len(cityList)):

        fromCity = cityList[fromCityIndex]
        fromCity = fromCity.split('\r')[0].decode('utf-8')
        if startCityFind:
            pass
        else:

            if fromCity == crashSave['fromCity']:
                startCityFind = True

            else:
                continue


        result = fromCity + "\t:"

        for toCityIndex in range(len(cityList)):
            toCity = cityList[toCityIndex]
            toCity = toCity.split('\r')[0].decode('utf-8')
            print toCity
            if toCityFind:
                pass
            else:
                if toCity == crashSave['toCity']:
                    toCityFind = True
                else:
                    continue
            initMap()
            if fromCity == toCity:
                number = '/'
            else:
                if fromCity == '莱芜':
                    number = 0
                    for fromCity2 in ['莱芜东', '莱芜西']:
                        fromStationCode = getStationByName(fromCity2)
                        toStationCode = getStationByName(toCity)
                        if toStationCode == 0:
                            number = '/'
                            break
                        else:
                            trains = queryTickets(queryDate, fromStationCode, toStationCode)
                            if trains == RET_ERR or trains == RET_OK:
                                break
                            number += len(trains['data'])
                            countByTime(trains['data'])
                elif toCity == '莱芜':
                    number = 0
                    for toCity2 in ['莱芜东', '莱芜西']:
                        fromStationCode = getStationByName(fromCity)
                        toStationCode = getStationByName(toCity2)
                        if fromStationCode == 0:
                            number = '/'
                            break
                        else:
                            trains = queryTickets(queryDate, fromStationCode, toStationCode)
                            if trains == RET_ERR or trains == RET_OK:
                                break
                            number += len(trains['data'])
                            countByTime(trains['data'])
                else:
                    fromStationCode = getStationByName(fromCity)
                    toStationCode = getStationByName(toCity)
                    print fromStationCode
                    if fromStationCode == 0 or toStationCode == 0:
                        number = '/'
                    else:
                        trains = queryTickets(queryDate, fromStationCode, toStationCode)
                        if trains == RET_ERR or trains == RET_OK:
                            break
                        number = len(trains['data'])
                        countByTime(trains['data'])

            print (fromCity + "->" + toCity + ":" + str(number))
            result = result + str(number) + ','
            print startTimeMap
            print arriveTimeMap
            for hour in range(EXCEL_SHEET_NUMBER):
                startSheet = startWB.get_sheet_by_name(str(hour))
                arriveSheet = arriveWB.get_sheet_by_name(str(hour))
	        sumSheet = sumWB.get_sheet_by_name("sum")
                if number == '/':
                    startSheet.cell(row=fromCityIndex + 2, column=toCityIndex + 2, value='/')
                    arriveSheet.cell(row=fromCityIndex + 2, column=toCityIndex + 2, value='/')
                    sumSheet.cell(row=fromCityIndex + 2, column=toCityIndex + 2, value='/')                    
                else:
                    startSheet.cell(row=fromCityIndex + 2, column=toCityIndex + 2, value=str(startTimeMap[hour]))
                    arriveSheet.cell(row=fromCityIndex + 2, column=toCityIndex + 2, value=str(arriveTimeMap[hour]))
                    sumSheet.cell(row=fromCityIndex + 2, column=toCityIndex + 2, value=str(number))                    

    startWB.save(filename=SAVE_PATH + START_EXCEL_NAME)
    arriveWB.save(filename=SAVE_PATH + ARRIVE_EXCEL_NAME)
    sumWB.save(filename=SAVE_PATH + SUM_EXCEL_NAME)

    print("*********all over*************")


def countByTime(data):
    for train in data:
        if train['queryLeftNewDTO']['controlled_train_flag'] == '0': # 过滤掉停运被控车次
            startHour = int(train['queryLeftNewDTO']['start_time'].split(':')[0])
            startTimeMap[startHour] += 1
            arriveHour = int(train['queryLeftNewDTO']['arrive_time'].split(':')[0])
            arriveTimeMap[arriveHour] += 1

# 判断start.xls和arrive.xls是否已存在于result目录下，存在则读取；不存在就创建
def initialWB():
    files = [f for f in os.listdir(SAVE_PATH) if os.path.isfile(SAVE_PATH + f)]
    startInitial = False
    arriveInitial = False
    sumInitial = False
    for f in files:
        if f == ARRIVE_EXCEL_NAME:
            global arriveWB
            arriveWB = load_workbook(filename=SAVE_PATH + ARRIVE_EXCEL_NAME)
            arriveInitial = True
        elif f == START_EXCEL_NAME:
            global startWB
            startWB = load_workbook(filename=SAVE_PATH + START_EXCEL_NAME)
            startInitial = True
        elif f == SUM_EXCEL_NAME:
            global sumWB
            sumWB = load_workbook(filename=SAVE_PATH + SUM_EXCEL_NAME)
            sumInitial = True
    if not startInitial:
        writeSheetCityName(startWB)
    if not arriveInitial:
        writeSheetCityName(arriveWB)
    if not sumInitial:
        ws = sumWB.create_sheet(title="sum")
        for i in range(len(cityList)):
            ws.cell(column=1, row=i + 2, value=cityList[i].split('\n')[0])
            ws.cell(column=i + 2, row=1, value=cityList[i].split('\n')[0])


# 在excel中每个sheet页第一行和第一列写城市名
def writeSheetCityName(wb):
    for index in range(0, EXCEL_SHEET_NUMBER):
        ws = wb.create_sheet(title=str(index))
        for i in range(len(cityList)):
            ws.cell(column=1, row=i + 2, value=cityList[i].split('\n')[0])
            ws.cell(column=i + 2, row=1, value=cityList[i].split('\n')[0])

def initMap():
    for index in range(EXCEL_SHEET_NUMBER):
        startTimeMap[index] = 0
        arriveTimeMap[index] = 0

if __name__ == '__main__':
    start()
